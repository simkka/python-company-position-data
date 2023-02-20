import openpyxl
from datetime import date
import pandas as pd
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

thick_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

def currency_exposure_all_funds(df_file, list_of_ccy):

    nav_table = df_file[['FUND','MKT_VAL_USD']].groupby('FUND').sum().reset_index()
    df_file = pd.merge(df_file,nav_table,on="FUND",how='left')
    df_file = df_file[df_file['NET_NAV'] >= 1e6]
    df_file = df_file[df_file['POSN_TYPE'] != 'Fwrd Fx (prop)']
    # df_file.drop(df_file[df_file['POSN_TYPE'] == 'Fwrd Fx (prop)'].index, inplace=True)
    df_file['Weight'] = df_file['MKT_VAL_USD_x'] / df_file['MKT_VAL_USD_y']
    df_file['Weight'] = df_file['Weight']*100
    df_f = df_file.groupby(['FUND', 'INST_CCY']).agg({'Weight': 'sum'}).reset_index()
    df_pivot = df_f.pivot(index='FUND', columns='INST_CCY', values='Weight')
    df_pivot = df_pivot.fillna(0)

    # rest list
    df_file2 = df_pivot[df_pivot.columns[~df_pivot.columns.isin(list_of_ccy)]]
    df_pivot['REST'] = df_file2.sum(axis=1)
    list_of_ccy.append('REST')
    df_pivot.drop(columns=df_pivot.columns[~df_pivot.columns.isin(list_of_ccy)], inplace=True)
    df_pivot = pd.concat([df_pivot], keys=['Exposure'], axis=1)
    df_pivot['Exposure'] = df_pivot['Exposure'].round(decimals=2)


    FilePath = "CurrencyExposure_" + str(today) + ".xlsx"
    writer = pd.ExcelWriter(FilePath, engine='openpyxl')
    df_pivot.to_excel(writer, sheet_name='All funds_exposure', startrow=3, startcol=0)
    writer.save()
    wb = openpyxl.load_workbook(FilePath)
    active = wb['All funds_exposure']
    active.cell(row=1, column=1).value = str(today)
    for column_cells in active.columns:
        active.column_dimensions[column_cells[0].column_letter].width = 15
    active.column_dimensions['A'].width = 30
    active['B4'].border = thick_border
    for row in range(4,active.max_row+1):
        cell=active.cell(row,1)
        cell.border = thick_border
    for row in range(1,active.max_row+1):
        for col in range(1, active.max_column + 1):
            cell=active.cell(row,col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(FilePath)

def currency_exposure_list(df_file, list_of_funds, list_of_ccy):
    df_file = df_file[df_file['FUND'].isin(list_of_funds)]
    nav_table = df_file[['FUND','MKT_VAL_USD']].groupby('FUND').sum().reset_index()
    df_file = pd.merge(df_file,nav_table,on="FUND",how='left')
    df_file = df_file[df_file['NET_NAV'] >= 1e6]
    df_file = df_file[df_file['POSN_TYPE'] != 'Fwrd Fx (prop)']
    # df_file.drop(df_file[df_file['POSN_TYPE'] == 'Fwrd Fx (prop)'].index, inplace=True)
    df_file['Weight'] = df_file['MKT_VAL_USD_x'] / df_file['MKT_VAL_USD_y']
    df_file['Weight'] = df_file['Weight']*100
    df_f = df_file.groupby(['FUND', 'INST_CCY']).agg({'Weight': 'sum'}).reset_index()
    df_pivot = df_f.pivot(index='FUND', columns='INST_CCY', values='Weight')
    df_pivot = df_pivot.fillna(0)

    # rest list
    df_file2 = df_pivot[df_pivot.columns[~df_pivot.columns.isin(list_of_ccy)]]
    df_pivot['REST'] = df_file2.sum(axis=1)
    list_of_ccy.append('REST')
    df_pivot.drop(columns=df_pivot.columns[~df_pivot.columns.isin(list_of_ccy)], inplace=True)
    df_pivot = pd.concat([df_pivot], keys=['Exposure'], axis=1)
    df_pivot['Exposure'] = df_pivot['Exposure'].round(decimals=2)

    FilePath = "CurrencyExposure_" + str(today) + ".xlsx"
    ExcelWorkbook = load_workbook(FilePath)
    writer = pd.ExcelWriter(FilePath, engine='openpyxl')
    writer.book = ExcelWorkbook
    df_pivot.to_excel(writer, sheet_name= 'list_exposure', startrow=3, startcol=0)
    writer.save()
    wb = openpyxl.load_workbook(FilePath)
    active = wb['list_exposure']
    active.cell(row=1, column=1).value = str(today)
    for column_cells in active.columns:
        active.column_dimensions[column_cells[0].column_letter].width = 15
    active.column_dimensions['A'].width = 30
    active['B4'].border = thick_border
    for row in range(4,active.max_row+1):
        cell=active.cell(row,1)
        cell.border = thick_border
    for col in range(1,active.max_column+1):
        cell=active.cell(4,col)
        cell.border = thick_border
    for row in range(1,active.max_row+1):
        for col in range(1, active.max_column + 1):
            cell=active.cell(row,col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(FilePath)

def send_email(me,recipients,FilePath ):
    msg = MIMEMultipart()
    msg["Subject"] = 'Quarterly Currency Exposure ' + str(today)
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(FilePath, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename=%s' % FilePath)
    msg.attach(part)

    server = smtplib.SMTP('mailserver', 25)
    server.sendmail(me, recipients, msg=msg.as_string())
    server.quit()

File= 'TT Position Data 7th July 22 updated.xlsx'
df_file = pd.read_excel(File)
list_of_funds = ['BICGEM', 'IOOFEM']
list_of_ccy=['GBP', 'JPY', 'EUR', 'CHF', 'NOK', 'SEK', 'AUD', 'CAD', 'KRW', 'SGD', 'USD']
today = date.today()
me = "sinkevichm1@ttint.com"
recipients = "sinkevichm1@ttint.com"
FilePath = "CurrencyExposure_" + str(today) + ".xlsx"

currency_exposure_all_funds(df_file, list_of_ccy)
currency_exposure_list(df_file, list_of_funds, list_of_ccy)
# send_email(me,recipients,FilePath)